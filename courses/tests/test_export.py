from io import BytesIO

from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from courses.models import (
    CourseLessonContentStatus,
    CourseModuleContentStatus,
    CourseTask,
    CourseTaskContentStatus,
    CourseTaskInformationalType,
    CourseTaskKind,
    CourseTaskQuestionType,
    CourseTaskAnswerType,
    CourseTaskCheckType,
    ProgressStatus,
    UserCourseProgress,
    UserLessonProgress,
)
from courses.services.answers import TaskAnswerSubmitPayload, submit_user_task_answer
from courses.services.progress import recalculate_user_progresses_for_lesson

from .helpers import (
    create_choice_question_task,
    create_course,
    create_files_question_task,
    create_informational_task,
    create_lesson,
    create_module,
    create_staff_user,
    create_text_and_files_question_task,
    create_text_question_task,
    create_user,
    create_user_file,
)


class CourseResultsExportTests(TestCase):
    def setUp(self):
        self.admin_user = create_staff_user()
        self.client = Client()
        self.client.force_login(self.admin_user)

    def _read_workbook_rows(self, response):
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        return list(worksheet.iter_rows(values_only=True))

    def _export_rows(self, course):
        response = self.client.get(reverse("admin:courses_export_results", args=[course.id]))
        self.assertEqual(response.status_code, 200)
        return self._read_workbook_rows(response)

    def _stage_from_export(self, course):
        rows = self._export_rows(course)
        self.assertEqual(len(rows), 2)
        return rows[1][5]

    def _mark_course_started(self, user, course, *, percent: int = 1):
        return UserCourseProgress.objects.create(
            user=user,
            course=course,
            status=ProgressStatus.IN_PROGRESS,
            percent=percent,
        )

    def test_course_admin_export_contains_dynamic_task_columns_and_answers(self):
        course = create_course(title="Экспорт SQL")
        module = create_module(course, title="SQL basics", order=1)
        lesson = create_lesson(module, title="SELECT", order=1)
        student = create_user(prefix="export-student")
        create_user(prefix="export-not-started")

        info_task = create_informational_task(lesson, title="Введение", order=1)
        text_task = create_text_question_task(lesson, title="Опишите юнит-экономику", order=2)
        choice_task, options = create_choice_question_task(
            lesson,
            title="Какие метрики важны",
            order=3,
            answer_type="multiple_choice",
            options=[
                ("Retention", True),
                ("LTV", True),
                ("Bounce Rate", False),
            ],
        )
        spreadsheet = create_user_file(
            student,
            name="economics-model",
            extension="xlsx",
            mime_type=(
                "application/"
                "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        files_task = create_files_question_task(
            lesson,
            attachment_file=spreadsheet,
            title="Приложите расчёт",
            order=4,
        )
        create_text_and_files_question_task(
            lesson,
            title="Финальный вывод",
            order=5,
        )

        submit_user_task_answer(student, info_task, TaskAnswerSubmitPayload())
        recalculate_user_progresses_for_lesson(student, lesson)
        submit_user_task_answer(
            student,
            text_task,
            TaskAnswerSubmitPayload(answer_text="Указал юнит-экономику."),
        )
        recalculate_user_progresses_for_lesson(student, lesson)
        submit_user_task_answer(
            student,
            choice_task,
            TaskAnswerSubmitPayload(option_ids=[options[0].id, options[1].id]),
        )
        recalculate_user_progresses_for_lesson(student, lesson)
        submit_user_task_answer(
            student,
            files_task,
            TaskAnswerSubmitPayload(file_ids=[spreadsheet.pk]),
        )
        recalculate_user_progresses_for_lesson(student, lesson)

        response = self.client.get(
            reverse("admin:courses_export_results", args=[course.id])
        )
        rows = self._read_workbook_rows(response)

        self.assertEqual(response.status_code, 200)
        self.assertIn(".xlsx", response["Content-Disposition"])
        self.assertEqual(len(rows), 2)

        header = rows[0]
        data_row = rows[1]

        self.assertEqual(
            header[:6],
            (
                "Имя и Фамилия",
                "Email",
                "Прогресс курса, %",
                "Дата начала прохождения",
                "Название курса",
                "Текущий этап",
            ),
        )
        self.assertNotIn("Введение", "\n".join(str(value) for value in header))
        self.assertIn(
            "Модуль 1: SQL basics / Урок 1: SELECT / Задание 2: Опишите юнит-экономику",
            header,
        )
        self.assertIn(
            "Модуль 1: SQL basics / Урок 1: SELECT / Задание 5: Финальный вывод",
            header,
        )

        self.assertEqual(data_row[0], f"{student.first_name} {student.last_name}")
        self.assertEqual(data_row[1], student.email)
        self.assertEqual(data_row[2], 80)
        self.assertEqual(data_row[4], course.title)
        self.assertEqual(
            data_row[6:10],
            (
                "Указал юнит-экономику.",
                "Retention\nLTV",
                spreadsheet.link,
                None,
            ),
        )
        self.assertEqual(
            data_row[5],
            "Модуль 1: SQL basics / Урок 1: SELECT / Задание 5: Финальный вывод",
        )

    def test_course_admin_export_marks_completed_course_stage(self):
        course = create_course(title="Экспорт аналитики")
        module = create_module(course, title="Метрики", order=1)
        lesson = create_lesson(module, title="Базовые метрики", order=1)
        student = create_user(prefix="export-completed")
        task = create_text_question_task(
            lesson,
            title="Что такое retention",
            order=1,
        )

        submit_user_task_answer(
            student,
            task,
            TaskAnswerSubmitPayload(answer_text="Retention отражает возврат пользователей."),
        )
        recalculate_user_progresses_for_lesson(student, lesson)

        response = self.client.get(
            reverse("admin:courses_export_results", args=[course.id])
        )
        rows = self._read_workbook_rows(response)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][2], 100)
        self.assertEqual(rows[1][5], "Курс завершён")
        self.assertEqual(rows[1][6], "Retention отражает возврат пользователей.")

    def test_course_admin_export_resolves_next_stage_for_not_started_next_module(self):
        course = create_course(title="Экспорт этапа")
        module_one = create_module(course, title="Основы", order=1)
        module_two = create_module(course, title="Практика", order=2)
        lesson_one = create_lesson(module_one, title="Денежный поток", order=1)
        lesson_two = create_lesson(module_one, title="Runway", order=2)
        lesson_three = create_lesson(module_two, title="Финальный блок", order=1)
        student = create_user(prefix="export-next-stage")

        module_one_tasks = [
            create_informational_task(lesson_one, title="Введение в cash flow", order=1),
            create_text_question_task(lesson_one, title="Что такое cash flow", order=2),
            create_informational_task(lesson_two, title="Что такое runway", order=1),
            create_text_question_task(lesson_two, title="Как считать runway", order=2),
        ]
        module_two_first_task = create_informational_task(
            lesson_three,
            title="Финальная инструкция",
            order=1,
        )

        submit_user_task_answer(student, module_one_tasks[0], TaskAnswerSubmitPayload())
        submit_user_task_answer(
            student,
            module_one_tasks[1],
            TaskAnswerSubmitPayload(answer_text="Cash flow показывает движение денег."),
        )
        recalculate_user_progresses_for_lesson(student, lesson_one)

        submit_user_task_answer(student, module_one_tasks[2], TaskAnswerSubmitPayload())
        submit_user_task_answer(
            student,
            module_one_tasks[3],
            TaskAnswerSubmitPayload(answer_text="Runway считают по burn rate."),
        )
        recalculate_user_progresses_for_lesson(student, lesson_two)

        response = self.client.get(
            reverse("admin:courses_export_results", args=[course.id])
        )
        rows = self._read_workbook_rows(response)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 2)
        self.assertEqual(
            rows[1][5],
            (
                f"Модуль 2: Практика / Урок 1: Финальный блок / "
                f"Задание 1: {module_two_first_task.title}"
            ),
        )

    def test_course_admin_export_resolves_next_stage_for_not_started_next_lesson(self):
        course = create_course(title="Экспорт следующего урока")
        module = create_module(course, title="Основной модуль", order=1)
        lesson_one = create_lesson(module, title="Урок 1", order=1)
        lesson_two = create_lesson(module, title="Урок 2", order=2)
        student = create_user(prefix="export-next-lesson")

        info_task = create_informational_task(lesson_one, title="Введение", order=1)
        create_text_question_task(lesson_one, title="Ответ по первому уроку", order=2)
        next_task = create_informational_task(lesson_two, title="Старт второго урока", order=1)

        submit_user_task_answer(student, info_task, TaskAnswerSubmitPayload())
        submit_user_task_answer(
            student,
            lesson_one.tasks.get(order=2),
            TaskAnswerSubmitPayload(answer_text="Первый урок завершён."),
        )
        recalculate_user_progresses_for_lesson(student, lesson_one)

        self.assertEqual(
            self._stage_from_export(course),
            f"Модуль 1: Основной модуль / Урок 2: Урок 2 / Задание 1: {next_task.title}",
        )

    def test_course_admin_export_uses_first_published_task_when_only_course_progress_exists(self):
        course = create_course(title="Экспорт старта без уроков")
        module = create_module(course, title="Структура", order=1)
        lesson = create_lesson(module, title="Первый урок", order=1)
        first_task = create_informational_task(lesson, title="Первое действие", order=1)
        create_text_question_task(lesson, title="Следующий вопрос", order=2)
        student = create_user(prefix="export-only-course-progress")

        self._mark_course_started(student, course)

        self.assertEqual(
            self._stage_from_export(course),
            f"Модуль 1: Структура / Урок 1: Первый урок / Задание 1: {first_task.title}",
        )

    def test_course_admin_export_uses_lesson_header_when_first_unfinished_lesson_has_no_tasks(self):
        course = create_course(title="Экспорт урока без заданий")
        module = create_module(course, title="Только теория", order=1)
        create_lesson(module, title="Пустой урок", order=1)
        student = create_user(prefix="export-empty-lesson")

        self._mark_course_started(student, course)

        self.assertEqual(
            self._stage_from_export(course),
            "Модуль 1: Только теория / Урок 1: Пустой урок",
        )

    def test_course_admin_export_uses_lesson_header_for_in_progress_lesson_without_current_task(self):
        course = create_course(title="Экспорт in_progress без задания")
        module = create_module(course, title="Модуль", order=1)
        lesson = create_lesson(module, title="Урок", order=1)
        student = create_user(prefix="export-in-progress-lesson")

        self._mark_course_started(student, course, percent=50)
        UserLessonProgress.objects.create(
            user=student,
            lesson=lesson,
            status=ProgressStatus.IN_PROGRESS,
            percent=50,
            current_task=None,
        )

        self.assertEqual(
            self._stage_from_export(course),
            "Модуль 1: Модуль / Урок 1: Урок",
        )

    def test_course_admin_export_ignores_unpublished_current_task_and_falls_back_to_lesson(self):
        course = create_course(title="Экспорт draft current task")
        module = create_module(course, title="Модуль", order=1)
        lesson = create_lesson(module, title="Урок", order=1)
        student = create_user(prefix="export-draft-current-task")

        draft_task = CourseTask.objects.create(
            lesson=lesson,
            title="Черновик",
            status=CourseTaskContentStatus.DRAFT,
            task_kind=CourseTaskKind.INFORMATIONAL,
            informational_type=CourseTaskInformationalType.TEXT,
            body_text="Черновик",
            order=1,
        )

        self._mark_course_started(student, course, percent=30)
        UserLessonProgress.objects.create(
            user=student,
            lesson=lesson,
            status=ProgressStatus.IN_PROGRESS,
            percent=30,
            current_task=draft_task,
        )

        self.assertEqual(
            self._stage_from_export(course),
            "Модуль 1: Модуль / Урок 1: Урок",
        )

    def test_course_admin_export_is_undefined_when_all_published_lessons_are_completed_but_course_is_not(self):
        course = create_course(title="Экспорт broken completed lessons")
        module = create_module(course, title="Модуль", order=1)
        lesson = create_lesson(module, title="Урок", order=1)
        student = create_user(prefix="export-broken-completed")

        self._mark_course_started(student, course, percent=99)
        UserLessonProgress.objects.create(
            user=student,
            lesson=lesson,
            status=ProgressStatus.COMPLETED,
            percent=100,
        )

        self.assertEqual(self._stage_from_export(course), "Этап не определён")

    def test_course_admin_export_is_undefined_when_course_has_no_published_lessons(self):
        course = create_course(title="Экспорт без опубликованных уроков")
        draft_module = create_module(
            course,
            title="Черновой модуль",
            order=1,
            status=CourseModuleContentStatus.DRAFT,
        )
        create_lesson(
            draft_module,
            title="Черновой урок",
            order=1,
            status=CourseLessonContentStatus.DRAFT,
        )
        student = create_user(prefix="export-no-published-lessons")

        self._mark_course_started(student, course)

        self.assertEqual(self._stage_from_export(course), "Этап не определён")

    def test_course_admin_export_skips_draft_modules_in_stage_fallback(self):
        course = create_course(title="Экспорт пропуска draft модуля")
        draft_module = create_module(
            course,
            title="Черновик",
            order=1,
            status=CourseModuleContentStatus.DRAFT,
        )
        create_lesson(draft_module, title="Черновой урок", order=1)
        module = create_module(course, title="Публикация", order=2)
        lesson = create_lesson(module, title="Старт", order=1)
        first_task = create_text_question_task(lesson, title="Первый вопрос", order=1)
        student = create_user(prefix="export-skip-draft-module")

        self._mark_course_started(student, course)

        self.assertEqual(
            self._stage_from_export(course),
            f"Модуль 2: Публикация / Урок 1: Старт / Задание 1: {first_task.title}",
        )

    def test_course_admin_export_skips_draft_lessons_in_stage_fallback(self):
        course = create_course(title="Экспорт пропуска draft урока")
        module = create_module(course, title="Модуль", order=1)
        draft_lesson = create_lesson(
            module,
            title="Черновой урок",
            order=1,
            status=CourseLessonContentStatus.DRAFT,
        )
        CourseTask.objects.create(
            lesson=draft_lesson,
            title="Черновое задание",
            status=CourseTaskContentStatus.DRAFT,
            task_kind=CourseTaskKind.QUESTION,
            question_type=CourseTaskQuestionType.TEXT,
            answer_type=CourseTaskAnswerType.TEXT,
            check_type=CourseTaskCheckType.WITHOUT_REVIEW,
            body_text="Черновик",
            order=1,
        )
        published_lesson = create_lesson(module, title="Рабочий урок", order=2)
        first_task = create_text_question_task(
            published_lesson,
            title="Первое опубликованное задание",
            order=1,
        )
        student = create_user(prefix="export-skip-draft-lesson")

        self._mark_course_started(student, course)

        self.assertEqual(
            self._stage_from_export(course),
            f"Модуль 1: Модуль / Урок 2: Рабочий урок / Задание 1: {first_task.title}",
        )
