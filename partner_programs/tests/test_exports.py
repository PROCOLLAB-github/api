import io

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from partner_programs.models import PartnerProgramFieldValue
from partner_programs.services import (
    prepare_project_scores_export_data,
    row_dict_for_link,
)
from partner_programs.tests.helpers import (
    create_partner_program,
    create_program_field,
    create_program_project,
    create_project,
    create_user,
)
from project_rates.models import Criteria, ProjectScore
from projects.models import Collaborator


class PartnerProgramProjectScoresExportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.manager = create_user(prefix="program-rates-export-manager")
        self.program = create_partner_program(name="Rates Export Program")
        self.program.managers.add(self.manager)

    def test_manager_can_export_project_scores_to_xlsx(self):
        field = create_program_field(
            self.program,
            name="track",
            label="Track",
        )
        project = create_project(name="Rated project")
        program_project = create_program_project(self.program, project=project)
        PartnerProgramFieldValue.objects.create(
            program_project=program_project,
            field=field,
            value_text="ai",
        )
        score_criteria = Criteria.objects.create(
            partner_program=self.program,
            name="Score",
            type="str",
        )
        comment_criteria = Criteria.objects.get(
            partner_program=self.program,
            name="Комментарий",
        )
        expert = create_user(prefix="program-rates-expert", last_name="Ivanov")
        second_expert = create_user(
            prefix="program-rates-second-expert",
            last_name="Petrov",
        )
        ProjectScore.objects.create(
            criteria=score_criteria,
            user=expert,
            project=project,
            value="9",
        )
        ProjectScore.objects.create(
            criteria=comment_criteria,
            user=expert,
            project=project,
            value="Strong project",
        )
        ProjectScore.objects.create(
            criteria=score_criteria,
            user=second_expert,
            project=project,
            value="8",
        )
        self.client.force_authenticate(self.manager)

        response = self.client.get(f"/programs/{self.program.id}/export-rates/")

        self.assertEqual(response.status_code, 200)
        self.assertIn(".xlsx", response["Content-Disposition"])

        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        workbook.close()

        header = list(rows[0])
        row_by_expert = {
            dict(zip(header, row))["Фамилия эксперта"]: dict(zip(header, row))
            for row in rows[1:]
        }
        self.assertEqual(row_by_expert["Ivanov"]["Название проекта"], "Rated project")
        self.assertEqual(row_by_expert["Ivanov"]["Track"], "ai")
        self.assertEqual(row_by_expert["Ivanov"]["Score"], "9")
        self.assertEqual(row_by_expert["Ivanov"]["Комментарий"], "Strong project")
        self.assertEqual(row_by_expert["Petrov"]["Score"], "8")
        self.assertIsNone(row_by_expert["Petrov"]["Комментарий"])

    def test_scores_export_data_returns_empty_row_when_criteria_exist_without_scores(self):
        Criteria.objects.create(
            partner_program=self.program,
            name="Score",
            type="str",
        )
        create_program_field(
            self.program,
            name="track",
            label="Track",
        )

        export_data = prepare_project_scores_export_data(self.program.id)

        self.assertEqual(len(export_data), 1)
        self.assertEqual(export_data[0]["Название проекта"], "")
        self.assertEqual(export_data[0]["Фамилия эксперта"], "")
        self.assertEqual(export_data[0]["Track"], "")
        self.assertEqual(export_data[0]["Score"], "")
        self.assertEqual(export_data[0]["Комментарий"], "")

    def test_non_manager_cannot_export_project_scores(self):
        outsider = create_user(prefix="program-rates-export-outsider")
        self.client.force_authenticate(outsider)

        response = self.client.get(f"/programs/{self.program.id}/export-rates/")

        self.assertEqual(response.status_code, 403)

    def test_row_dict_for_link_contains_team_and_program_field_values(self):
        leader = create_user(
            prefix="program-export-leader",
            first_name="Leader",
            last_name="Owner",
        )
        member = create_user(
            prefix="program-export-member",
            first_name="Team",
            last_name="Member",
        )
        project = create_project(
            leader=leader,
            name="Team project",
            description="Exported project",
            region="Moscow",
            presentation_address="https://example.com/presentation",
        )
        Collaborator.objects.create(project=project, user=member, role="Developer")
        program_project = create_program_project(self.program, project=project)
        field = create_program_field(
            self.program,
            name="track",
            label="Track",
        )
        PartnerProgramFieldValue.objects.create(
            program_project=program_project,
            field=field,
            value_text="ai",
        )

        row = row_dict_for_link(
            program_project_link=program_project,
            extra_field_keys_order=["name:track"],
            row_number=1,
        )

        self.assertEqual(row["row_number"], 1)
        self.assertEqual(row["project_name"], "Team project")
        self.assertEqual(row["project_description"], "Exported project")
        self.assertEqual(row["project_region"], "Moscow")
        self.assertEqual(row["project_presentation"], "https://example.com/presentation")
        self.assertGreaterEqual(row["team_size"], 2)
        self.assertIn("Leader Owner", row["team_members"])
        self.assertIn("Team Member", row["team_members"])
        self.assertEqual(row["leader_full_name"], "Leader Owner")
        self.assertEqual(row["name:track"], "ai")
