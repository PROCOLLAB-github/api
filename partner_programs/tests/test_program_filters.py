import io

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from partner_programs.models import PartnerProgramFieldValue
from partner_programs.tests.helpers import (
    create_partner_program,
    create_program_field,
    create_program_project,
    create_project,
    create_user,
)
from users.models import CustomUser


class PartnerProgramProjectFilterAPITests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.manager = create_user(prefix="program-filter-manager")
        self.program = create_partner_program()
        self.program.managers.add(self.manager)

    def _create_expert(self, *, prefix, program):
        expert = create_user(prefix=prefix, user_type=CustomUser.EXPERT)
        expert.expert.programs.add(program)
        return expert

    def test_manager_can_get_filterable_program_fields(self):
        filterable = create_program_field(
            self.program,
            name="track",
            label="Track",
            field_type="select",
            options=["ai", "edu"],
            show_filter=True,
        )
        hidden = create_program_field(
            self.program,
            name="internal",
            label="Internal",
            show_filter=False,
        )
        self.client.force_authenticate(self.manager)

        response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(response.status_code, 200)
        field_ids = {item["id"] for item in response.data}
        self.assertIn(filterable.id, field_ids)
        self.assertNotIn(hidden.id, field_ids)

    def test_non_manager_cannot_get_filterable_program_fields(self):
        outsider = create_user(prefix="program-filter-outsider")
        self.client.force_authenticate(outsider)

        response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(response.status_code, 403)

    def test_anonymous_cannot_get_filterable_program_fields(self):
        response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(response.status_code, 401)

    def test_program_expert_can_get_filterable_program_fields(self):
        expert = self._create_expert(
            prefix="program-filter-expert",
            program=self.program,
        )
        self.client.force_authenticate(expert)

        response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(response.status_code, 200)

    def test_other_program_expert_cannot_get_filterable_program_fields(self):
        other_program = create_partner_program()
        expert = self._create_expert(
            prefix="other-program-filter-expert",
            program=other_program,
        )
        self.client.force_authenticate(expert)

        response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(response.status_code, 403)

    def test_staff_can_get_filterable_program_fields(self):
        staff = create_user(prefix="program-filter-staff", is_staff=True)
        self.client.force_authenticate(staff)

        response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(response.status_code, 200)

    def test_superuser_can_get_filterable_program_fields(self):
        superuser = create_user(
            prefix="program-filter-superuser",
            is_staff=True,
            is_superuser=True,
        )
        self.client.force_authenticate(superuser)

        response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(response.status_code, 200)

    def test_expert_receives_same_filter_schema_as_manager(self):
        create_program_field(
            self.program,
            name="track",
            label="Track",
            field_type="select",
            options=["ai", "edu"],
            show_filter=True,
        )
        expert = self._create_expert(
            prefix="program-filter-schema-expert",
            program=self.program,
        )

        self.client.force_authenticate(self.manager)
        manager_response = self.client.get(f"/programs/{self.program.id}/filters/")
        self.client.force_authenticate(expert)
        expert_response = self.client.get(f"/programs/{self.program.id}/filters/")

        self.assertEqual(manager_response.status_code, 200)
        self.assertEqual(expert_response.status_code, 200)
        self.assertEqual(expert_response.data, manager_response.data)

    def test_filter_schema_endpoint_remains_read_only(self):
        expert = self._create_expert(
            prefix="program-filter-read-only-expert",
            program=self.program,
        )
        self.client.force_authenticate(expert)

        response = self.client.post(
            f"/programs/{self.program.id}/filters/",
            {},
            format="json",
        )

        self.assertEqual(response.status_code, 405)

    def test_manager_can_filter_program_projects_by_field_value(self):
        field = create_program_field(
            self.program,
            name="track",
            field_type="select",
            options=["ai", "edu"],
            show_filter=True,
        )
        matching_project = create_project(name="AI project")
        other_project = create_project(name="Education project")
        matching_link = create_program_project(self.program, project=matching_project)
        other_link = create_program_project(self.program, project=other_project)
        PartnerProgramFieldValue.objects.create(
            program_project=matching_link,
            field=field,
            value_text="ai",
        )
        PartnerProgramFieldValue.objects.create(
            program_project=other_link,
            field=field,
            value_text="edu",
        )
        self.client.force_authenticate(self.manager)

        response = self.client.post(
            f"/programs/{self.program.id}/projects/filter/",
            {"filters": {"track": ["ai"]}},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        project_ids = {item["id"] for item in response.data["results"]}
        self.assertEqual(project_ids, {matching_project.id})

    def test_filter_rejects_field_that_is_not_filterable(self):
        create_program_field(
            self.program,
            name="internal",
            field_type="select",
            options=["yes", "no"],
            show_filter=False,
        )
        self.client.force_authenticate(self.manager)

        response = self.client.post(
            f"/programs/{self.program.id}/projects/filter/",
            {"filters": {"internal": ["yes"]}},
            format="json",
        )

        self.assertEqual(response.status_code, 400)

    def test_filter_rejects_unknown_program_field(self):
        self.client.force_authenticate(self.manager)

        response = self.client.post(
            f"/programs/{self.program.id}/projects/filter/",
            {"filters": {"missing": ["yes"]}},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("Поля не найденные", response.data["detail"])

    def test_filter_rejects_value_outside_field_options(self):
        create_program_field(
            self.program,
            name="track",
            field_type="select",
            options=["ai", "edu"],
            show_filter=True,
        )
        self.client.force_authenticate(self.manager)

        response = self.client.post(
            f"/programs/{self.program.id}/projects/filter/",
            {"filters": {"track": ["wrong"]}},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["invalid"], ["wrong"])

    def test_filter_rejects_filterable_field_without_options(self):
        create_program_field(
            self.program,
            name="comment",
            field_type="text",
            show_filter=True,
        )
        self.client.force_authenticate(self.manager)

        response = self.client.post(
            f"/programs/{self.program.id}/projects/filter/",
            {"filters": {"comment": ["value"]}},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("не имеет вариантов", response.data["detail"])


class PartnerProgramProjectsAPITests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.manager = create_user(prefix="program-projects-manager")
        self.program = create_partner_program()
        self.program.managers.add(self.manager)

    def test_manager_can_get_program_projects(self):
        project = create_project(name="Program project")
        create_program_project(self.program, project=project)
        self.client.force_authenticate(self.manager)

        response = self.client.get(f"/programs/{self.program.id}/projects/")

        self.assertEqual(response.status_code, 200)
        project_ids = {item["id"] for item in response.data["results"]}
        self.assertEqual(project_ids, {project.id})

    def test_non_manager_cannot_get_program_projects(self):
        outsider = create_user(prefix="program-projects-outsider")
        self.client.force_authenticate(outsider)

        response = self.client.get(f"/programs/{self.program.id}/projects/")

        self.assertEqual(response.status_code, 403)


class PartnerProgramProjectExportAPITests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.manager = create_user(prefix="program-export-manager")
        self.program = create_partner_program(name="Export Program")
        self.program.managers.add(self.manager)

    def test_manager_can_export_program_projects_to_xlsx(self):
        field = create_program_field(
            self.program,
            name="track",
            label="Track",
        )
        submitted_project = create_project(name="Submitted project")
        draft_project = create_project(name="Draft project")
        submitted_link = create_program_project(
            self.program,
            project=submitted_project,
            submitted=True,
        )
        draft_link = create_program_project(
            self.program,
            project=draft_project,
            submitted=False,
        )
        PartnerProgramFieldValue.objects.create(
            program_project=submitted_link,
            field=field,
            value_text="ai",
        )
        PartnerProgramFieldValue.objects.create(
            program_project=draft_link,
            field=field,
            value_text="edu",
        )
        self.client.force_authenticate(self.manager)

        response = self.client.get(
            f"/programs/{self.program.id}/export-projects/",
            {"only_submitted": "true"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(".xlsx", response["Content-Disposition"])

        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        rows = list(workbook["Проекты"].iter_rows(values_only=True))
        workbook.close()

        header = rows[0]
        project_names = [row[1] for row in rows[1:]]
        self.assertIn("Track", header)
        self.assertEqual(project_names, ["Submitted project"])

    def test_non_manager_cannot_export_program_projects(self):
        outsider = create_user(prefix="program-export-outsider")
        self.client.force_authenticate(outsider)

        response = self.client.get(f"/programs/{self.program.id}/export-projects/")

        self.assertEqual(response.status_code, 403)
